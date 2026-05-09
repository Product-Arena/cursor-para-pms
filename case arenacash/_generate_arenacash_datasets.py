"""
One-off generator for ArenaCash teaching datasets (weekly KPIs, weekly support rollup, financeiro xlsx).
Run from case folder: .venv/bin/python _generate_arenacash_datasets.py
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

CASE_DIR = Path(__file__).resolve().parent


def _thursday_of_week(week_monday: date) -> date:
    return week_monday + timedelta(days=3)


def month_week_starts(year: int, month: int) -> list[date]:
    """ISO weeks attributed to a calendar month via that week's Thursday (no double borders)."""
    if month == 12:
        last = date(year, 12, 31)
    else:
        last = date(year, month + 1, 1) - timedelta(days=1)
    first = date(year, month, 1)
    mondays: list[date] = []
    d = first
    while d <= last:
        mon = d - timedelta(days=d.weekday())
        th = _thursday_of_week(mon)
        if th.year == year and th.month == month and mon not in mondays:
            mondays.append(mon)
        d += timedelta(days=1)
    return sorted(set(mondays))


def weights_for_weeks(n: int) -> list[float]:
    if n <= 0:
        return []
    if n <= 4:
        base = [0.24, 0.26, 0.25, 0.25]
        return base[:n]
    if n == 5:
        return [0.19, 0.21, 0.20, 0.20, 0.20]
    return [1.0 / n] * n


def split_monthly_across_weeks(value: float, n_weeks: int) -> list[float]:
    w = weights_for_weeks(n_weeks)
    s = sum(w)
    return [value * (wi / s) for wi in w]


def build_weekly_kpis() -> pd.DataFrame:
    monthly = pd.read_csv(CASE_DIR / "dados-quarter-q1.csv")
    rows: list[dict] = []

    for _, r in monthly.iterrows():
        y, m = map(int, str(r["mes"]).split("-"))
        week_starts = month_week_starts(y, m)
        n = len(week_starts)

        def spl(col: str) -> list[float]:
            return split_monthly_across_weeks(float(r[col]), n)

        ns = spl("novos_signups")
        aa = spl("ativados_30d")
        pa = spl("pagantes_ativos_fim_mes")  # end-of-month stock; weekly is "snapshot proxy"
        cx = spl("cancelamentos_mes")
        ts = spl("tickets_suporte_abertos")
        mau = spl("mau")
        dau = spl("dau_medio")
        aum = spl("aum_total_brl_milhoes")

        for i, ws in enumerate(week_starts):
            iso = ws.isocalendar()
            week_label = f"{iso.year}-W{iso.week:02d}"
            taxa = (aa[i] / ns[i] * 100) if ns[i] else 0.0
            churn_w = (cx[i] / pa[i] * 100) if pa[i] else 0.0
            rows.append(
                {
                    "week_start_date": ws.isoformat(),
                    "year_iso": iso.year,
                    "week_iso": iso.week,
                    "week_label": week_label,
                    "mes_calendario": r["mes"],
                    "plano": r["plano"],
                    "novos_signups": round(ns[i]),
                    "ativados_30d": round(aa[i]),
                    "taxa_ativacao_pct": round(taxa, 1),
                    "pagantes_ativos_media_semana": round(pa[i]),
                    "cancelamentos_semana": round(cx[i]),
                    "churn_semanal_pct": round(churn_w, 2),
                    "nps": int(r["nps"]),
                    "mau": round(mau[i]),
                    "dau_medio": round(dau[i]),
                    "ticket_medio_primeira_aplicacao_brl": int(r["ticket_medio_primeira_aplicacao_brl"]),
                    "aum_total_brl_milhoes": round(aum[i], 1),
                    "tickets_suporte_abertos": round(ts[i]),
                }
            )

    out = pd.DataFrame(rows)
    out = out.sort_values(["week_start_date", "plano"]).reset_index(drop=True)
    return out


def build_support_weekly() -> pd.DataFrame:
    att = pd.read_csv(CASE_DIR / "atendimento-q1.csv")
    att["data_abertura"] = pd.to_datetime(att["data_abertura"])
    att["week_start"] = att["data_abertura"].dt.normalize() - pd.to_timedelta(
        att["data_abertura"].dt.weekday, unit="D"
    )
    iso = att["week_start"].dt.isocalendar()
    att["year_iso"] = iso.year.astype(int)
    att["week_iso"] = iso.week.astype(int)

    def pct_escalado(group: pd.DataFrame) -> float:
        return float((group["status"] == "escalado").mean() * 100)

    def pct_critica(group: pd.DataFrame) -> float:
        return float((group["prioridade"] == "critica").mean() * 100)

    esc = att.groupby("week_start", group_keys=False).apply(pct_escalado, include_groups=False)
    crit = att.groupby("week_start", group_keys=False).apply(pct_critica, include_groups=False)

    g = att.groupby(["week_start"], as_index=False).agg(
        tickets_abertos=("ticket_id", "count"),
        tempo_resolucao_medio_horas=("tempo_resolucao_horas", "mean"),
        csat_medio=("csat_nota", "mean"),
        year_iso=("year_iso", "first"),
        week_iso=("week_iso", "first"),
    )
    g["week_label"] = g.apply(
        lambda r: f"{int(r['year_iso'])}-W{int(r['week_iso']):02d}", axis=1
    )
    g["pct_escalados"] = g["week_start"].map(esc)
    g["pct_critica_prioridade"] = g["week_start"].map(crit)
    g["tempo_resolucao_medio_horas"] = g["tempo_resolucao_medio_horas"].round(2)
    g["csat_medio"] = g["csat_medio"].round(2)
    g["pct_escalados"] = g["pct_escalados"].round(2)
    g["pct_critica_prioridade"] = g["pct_critica_prioridade"].round(2)
    g.rename(columns={"week_start": "week_start_date"}, inplace=True)
    g["week_start_date"] = pd.to_datetime(g["week_start_date"]).dt.date.astype(str)
    return g[
        [
            "week_start_date",
            "week_label",
            "tickets_abertos",
            "tempo_resolucao_medio_horas",
            "csat_medio",
            "pct_escalados",
            "pct_critica_prioridade",
        ]
    ]


def build_financeiro_xlsx() -> None:
    # Monthly financials Jan-Mar 2026 (Q1 focus) + trailing quarter for context
    monthly_fin = pd.DataFrame(
        {
            "mes": [
                "2025-10",
                "2025-11",
                "2025-12",
                "2026-01",
                "2026-02",
                "2026-03",
            ],
            "receita_assinaturas_pro_brl_mil": [7.2, 7.8, 8.4, 9.1, 9.8, 10.2],
            "receita_spread_e_interchange_brl_mil": [18.5, 19.2, 20.1, 20.8, 21.5, 22.0],
            "custos_infra_e_fraud_brl_mil": [4.1, 4.2, 4.4, 4.8, 5.1, 5.4],
            "marketing_brl_mil": [12.0, 13.5, 14.2, 15.8, 18.2, 21.0],
            "people_cs_e_ops_brl_mil": [9.5, 9.7, 9.9, 10.2, 10.6, 11.0],
            "ebitda_brl_mil": [0.1, 0.6, 1.0, 0.6, -2.6, -5.2],
        }
    )

    weekly_rev = pd.DataFrame(
        {
            "week_start_date": pd.date_range(date(2026, 1, 6), periods=13, freq="W-MON"),
            "receita_total_brl_mil": [
                6.8,
                6.9,
                7.0,
                7.2,
                7.1,
                7.3,
                7.5,
                7.4,
                7.6,
                7.8,
                7.9,
                8.0,
                8.1,
            ],
            "custos_variaveis_brl_mil": [
                4.9,
                5.0,
                5.1,
                5.3,
                5.8,
                6.2,
                6.5,
                6.7,
                6.9,
                7.1,
                7.3,
                7.4,
                7.6,
            ],
        }
    )
    weekly_rev["week_start_date"] = weekly_rev["week_start_date"].dt.date.astype(str)
    weekly_rev["contribuicao_semanal_brl_mil"] = (
        weekly_rev["receita_total_brl_mil"] - weekly_rev["custos_variaveis_brl_mil"]
    ).round(2)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "monthly_financials"
    for r in dataframe_to_rows(monthly_fin, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("weekly_revenue_q1")
    for r in dataframe_to_rows(weekly_rev, index=False, header=True):
        ws2.append(r)

    ws3 = wb.create_sheet("notes")
    ws3.append(["note", "value"])
    ws3.append(["currency", "BRL thousands (mil R$) unless noted"])
    ws3.append(["fiscal_year", "calendar year"])
    ws3.append(["q1_scope", "2026-01 through 2026-03"])

    path = CASE_DIR / "financeiro-q1.xlsx"
    wb.save(path)
    print(f"Wrote {path}")


def main() -> None:
    weekly_kpis = build_weekly_kpis()
    out_kpis = CASE_DIR / "dados-semanais-arenacash.csv"
    weekly_kpis.to_csv(out_kpis, index=False)
    print(f"Wrote {out_kpis} ({len(weekly_kpis)} rows)")

    support_w = build_support_weekly()
    out_sup = CASE_DIR / "atendimento-resumo-semanal.csv"
    support_w.to_csv(out_sup, index=False)
    print(f"Wrote {out_sup} ({len(support_w)} rows)")

    build_financeiro_xlsx()


if __name__ == "__main__":
    main()
